"""
审计日志 API - 操作日志查询与管理
"""
from flask import Blueprint, request, jsonify, send_file
from app.services import AuditService
from app.decorators import require_permission

audit_bp = Blueprint('audit', __name__)


@audit_bp.route('/logs', methods=['GET'])
@require_permission('audit_view')
def get_audit_logs():
    """查询审计日志（管理员权限）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    filters = {
        'user_id': request.args.get('user_id', type=int),
        'action': request.args.get('action'),
        'resource_type': request.args.get('resource_type'),
        'status': request.args.get('status'),
        'username': request.args.get('username'),
        'start_time': request.args.get('start_time'),
        'end_time': request.args.get('end_time')
    }
    
    # 移除None值
    filters = {k: v for k, v in filters.items() if v is not None}
    
    result = AuditService.get_logs(filters=filters, page=page, per_page=per_page)
    return jsonify(result), 200


@audit_bp.route('/actions', methods=['GET'])
@require_permission('audit_view')
def get_action_types():
    """获取所有预定义的操作类型"""
    actions = [
        {'value': AuditService.LOGIN, 'label': '登录'},
        {'value': AuditService.LOGIN_FAILED, 'label': '登录失败'},
        {'value': AuditService.LOGOUT, 'label': '登出'},
        {'value': AuditService.USER_CREATE, 'label': '创建用户'},
        {'value': AuditService.USER_UPDATE, 'label': '更新用户'},
        {'value': AuditService.USER_DELETE, 'label': '删除用户'},
        {'value': AuditService.ROLE_CREATE, 'label': '创建角色'},
        {'value': AuditService.ROLE_UPDATE, 'label': '更新角色'},
        {'value': AuditService.ROLE_DELETE, 'label': '删除角色'},
        {'value': AuditService.ROLE_ASSIGN, 'label': '分配角色'},
        {'value': AuditService.PERMISSION_DENIED, 'label': '权限拒绝'},
        {'value': AuditService.PROJECT_CREATE, 'label': '创建项目'},
        {'value': AuditService.PROJECT_UPDATE, 'label': '更新项目'},
        {'value': AuditService.PROJECT_DELETE, 'label': '删除项目'},
        {'value': AuditService.TASK_CREATE, 'label': '创建任务'},
        {'value': AuditService.TASK_UPDATE, 'label': '更新任务'},
        {'value': AuditService.TASK_DELETE, 'label': '删除任务'},
        {'value': AuditService.APPROVAL_CREATE, 'label': '创建审批'},
        {'value': AuditService.APPROVAL_PROCESS, 'label': '处理审批'},
        {'value': AuditService.DATA_EXPORT, 'label': '数据导出'},
        {'value': AuditService.PASSWORD_CHANGE, 'label': '修改密码'},
        {'value': AuditService.TOKEN_REFRESH, 'label': 'Token刷新'}
    ]
    return jsonify({'actions': actions}), 200


@audit_bp.route('/stats', methods=['GET'])
@require_permission('audit_view')
def get_audit_stats():
    """获取审计统计概览"""
    from app.models import AuditLog
    from sqlalchemy import func
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    
    # 今日操作数
    today_start = datetime.combine(today, datetime.min.time())
    today_count = AuditLog.query.filter(AuditLog.created_at >= today_start).count()
    
    # 本周操作数
    week_start = datetime.combine(week_ago, datetime.min.time())
    week_count = AuditLog.query.filter(AuditLog.created_at >= week_start).count()
    
    # 登录失败数（安全指标）
    failed_logins = AuditLog.query.filter(
        AuditLog.action == AuditService.LOGIN_FAILED,
        AuditLog.created_at >= week_start
    ).count()
    
    # 权限拒绝数（安全指标）
    denied_count = AuditLog.query.filter(
        AuditLog.action == AuditService.PERMISSION_DENIED,
        AuditLog.created_at >= week_start
    ).count()
    
    # 操作类型分布（Top 10）
    action_dist = AuditLog.query.with_entities(
        AuditLog.action,
        func.count(AuditLog.id).label('count')
    ).filter(AuditLog.created_at >= week_start).group_by(AuditLog.action).order_by(func.count(AuditLog.id).desc()).limit(10).all()
    
    return jsonify({
        'today_count': today_count,
        'week_count': week_count,
        'failed_logins': failed_logins,
        'permission_denied': denied_count,
        'action_distribution': [{'action': a, 'count': c} for a, c in action_dist]
    }), 200


@audit_bp.route('/my-logs', methods=['GET'])
@require_permission('audit_view')
def get_my_logs():
    """获取当前用户的操作日志"""
    from flask_jwt_extended import get_jwt_identity
    user_id = get_jwt_identity()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    filters = {'user_id': user_id}
    result = AuditService.get_logs(filters=filters, page=page, per_page=per_page)
    return jsonify(result), 200


@audit_bp.route('/detail/<int:log_id>', methods=['GET'])
@require_permission('audit_view')
def get_audit_detail(log_id):
    """获取审计日志详情"""
    from app.models import AuditLog, User
    
    log = AuditLog.query.get(log_id)
    if not log:
        return jsonify({'message': '日志不存在'}), 404
    
    # 获取用户信息
    user = User.query.get(log.user_id) if log.user_id else None
    
    detail = {
        'id': log.id,
        'action': log.action,
        'resource_type': log.resource_type,
        'resource_id': log.resource_id,
        'detail': log.detail,
        'ip_address': log.ip_address,
        'status': log.status,
        'created_at': log.created_at.isoformat() if log.created_at else None,
        'user': {
            'id': user.id if user else None,
            'username': user.username if user else None,
            'real_name': user.real_name if user else None
        } if user else None
    }
    
    return jsonify({'log': detail}), 200


@audit_bp.route('/export', methods=['GET'])
@require_permission('audit_view')
def export_audit_logs():
    """导出审计日志为Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    # 获取筛选参数
    filters = {
        'user_id': request.args.get('user_id', type=int),
        'action': request.args.get('action'),
        'resource_type': request.args.get('resource_type'),
        'status': request.args.get('status'),
        'username': request.args.get('username'),
        'start_time': request.args.get('start_time'),
        'end_time': request.args.get('end_time')
    }
    filters = {k: v for k, v in filters.items() if v is not None}
    
    # 查询所有数据（不分页）
    result = AuditService.get_logs(filters=filters, page=1, per_page=10000)
    logs = result.get('logs', [])
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '审计日志'
    
    # 表头
    headers = ['时间', '操作人', '操作类型', '资源类型', '资源ID', 'IP地址', '状态', '详情']
    ws.append(headers)
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 数据行
    for log in logs:
        ws.append([
            log.get('created_at', ''),
            log.get('username', ''),
            log.get('action', ''),
            log.get('resource_type', ''),
            log.get('resource_id', ''),
            log.get('ip_address', ''),
            log.get('status', ''),
            str(log.get('detail', ''))[:500]
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 50
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
